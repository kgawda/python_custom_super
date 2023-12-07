import openpyxl

def create_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1)
    cell.value = "Hello World!"
    cell.hyperlink = "http://example.com"
    cell.font = openpyxl.styles.Font(color="0000FF", underline=openpyxl.styles.Font.UNDERLINE_SINGLE)
    ws.column_dimensions["A"].width = 40
    ws.merge_cells(start_row=1, start_column=2, end_row=4, end_column=4)

    wb.save("example_xlsx.xlsx")
    # PermissionError: [Errno 13] Permission denied: 'example_xlsx.xlsx' --> Excel z tym plikiem jest otwarty
    wb.close()

def load_file():
    wb = openpyxl.load_workbook("excel_example.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        # row jest tuplą - indeksowanie kolumn od 0!
        print(row[9].value, "at", row[9].coordinate)


if __name__ == "__main__":
    create_file()
    load_file()
